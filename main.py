import os
import openpyxl
import csv
import subprocess
import re

# Função para converter a segunda aba de um arquivo xlsx para csv
def xlsx_to_csv(xlsx_file, output_dir, delimiter='#'):
    try:
        # Carregar o arquivo xlsx
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        
        # Verificar se tem ao menos duas abas
        if len(wb.sheetnames) < 2:
            print(f"O arquivo {xlsx_file} não possui uma segunda aba.")
            return
        
        # Selecionar a segunda aba (index 1)
        sheet = wb[wb.sheetnames[1]]
        
        # Definir o nome do arquivo CSV base
        base_filename = "notas"
        
        # Extrair letras maiúsculas do nome do arquivo
        file_name = os.path.splitext(os.path.basename(xlsx_file))[0]  # Nome do arquivo sem extensão
        uppercase_parts = re.findall(r'[A-Z]+', file_name)  # Encontrar partes em maiúsculas
        
        # Concatenar as partes em maiúsculas ao nome base
        if uppercase_parts:
            base_filename += ''.join(uppercase_parts)  # Concatenar todas as partes em maiúsculas

        # Definir o nome do arquivo CSV
        csv_filename = f"{base_filename}.csv"
        
        # Definir o caminho completo do arquivo CSV
        csv_file = os.path.join(output_dir, csv_filename)
        
        # Escrever os dados da segunda aba em um arquivo CSV
        with open(csv_file, mode='w', newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)
        
        print(f"Convertido: {xlsx_file} -> {csv_file}")
    except Exception as e:
        print(f"Erro ao processar {xlsx_file}: {e}")

# Função para identificar o cliente a partir do nome do arquivo
def get_cliente_from_filename(filename):
    if 'arcelor' in filename.lower():
        return 'arcelor'
    elif 'tres' in filename.lower():
        return 'tres'
    elif 'lorenzetti' in filename.lower():
        return 'lorenzetti'
    elif 'mueller' in filename.lower():
        return 'mueller'
    elif 'muffato' in filename.lower():
        return 'muffato'
    elif 'pirahy' in filename.lower():
        return 'pirahy'
    else:
        return None

# Função para percorrer um diretório e converter todos os arquivos xlsx
def convert_all_xlsx_in_directory(input_dir, base_output_dir, delimiter='#'):
    # Percorrer todos os arquivos no diretório de entrada
    for root, _, files in os.walk(input_dir):
        for file in files:
            if file.endswith('.xlsx'):
                xlsx_file = os.path.join(root, file)
                
                # Identificar o cliente a partir do nome do arquivo
                cliente = get_cliente_from_filename(file)
                
                if cliente:
                    output_dir = os.path.join(base_output_dir, cliente)
                    
                    # Criar o diretório do cliente, se não existir
                    if not os.path.exists(output_dir):
                        os.makedirs(output_dir)
                    
                    # Converter e salvar no diretório correto
                    xlsx_to_csv(xlsx_file, output_dir, delimiter)
                else:
                    print(f"Cliente não identificado para o arquivo {file}. Verifique o nome.")
    
    # Executar o script enviarNotas com o argumento 6 ao final da conversão
    try:
        subprocess.run(["/usr/local/bin/enviarNotas", "7"], check=True)
        print("Script enviarNotas executado com sucesso com o argumento 7(Todos).")
    except subprocess.CalledProcessError as e:
        print(f"Erro ao executar o script enviarNotas: {e}")

# Diretório de entrada e base do diretório de saída
input_dir = '/home/sidney/Downloads'  # Diretório onde estão os arquivos xlsx
base_output_dir = '/home/sidney/Downloads'  # Diretório base onde estão as pastas dos clientes

# Executar a conversão
convert_all_xlsx_in_directory(input_dir, base_output_dir)
