import os
import yaml
import subprocess
import openpyxl
import csv
import re

# Função para carregar a configuração a partir do arquivo YAML
def carregar_configuracao():
    with open("clientes.yml", "r") as f:
        return yaml.safe_load(f)

# Função para extrair letras maiúsculas do nome do arquivo
def extrair_maiusculas(nome_arquivo):
    return ''.join(re.findall(r'[A-Z]', nome_arquivo))

# Função para formatar valores numéricos corretamente
def formatar_valor(valor):
    if isinstance(valor, float):
        if valor.is_integer():  # Se o número é inteiro, remover ".0"
            return str(int(valor))
    return str(valor)

# Função para converter a segunda aba de um arquivo xlsx para csv
def xlsx_to_csv(xlsx_file, output_dir, delimiter='#'):
    try:
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        if len(wb.sheetnames) < 2:
            print(f"O arquivo {xlsx_file} não possui uma segunda aba.")
            return None
        
        sheet = wb[wb.sheetnames[1]]
        
        # Extrair letras maiúsculas do nome do arquivo
        nome_base = "notas"
        maiusculas = extrair_maiusculas(os.path.basename(xlsx_file))

        # Concatenar as letras maiúsculas, se houver
        if maiusculas:
            nome_base += maiusculas

        # Nome do arquivo CSV convertido
        csv_filename = f"{nome_base}.csv"
        csv_file = os.path.join(output_dir, csv_filename)
        
        # Escrever os dados da segunda aba em um arquivo CSV
        with open(csv_file, mode='w', newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in sheet.iter_rows(values_only=True):
                # Usar a função formatar_valor para formatar os valores corretamente
                formatted_row = [formatar_valor(cell) for cell in row]
                writer.writerow(formatted_row)
        
        print(f"Convertido: {xlsx_file} -> {csv_file}")
        return csv_file
    except Exception as e:
        print(f"Erro ao processar {xlsx_file}: {e}")
        return None

# Função para enviar arquivo via SCP e rodar o script remoto
def enviar_arquivo(cliente_nome, arquivos):
    config = carregar_configuracao()

    cliente = config["clientes"].get(cliente_nome)
    if not cliente:
        print(f"Cliente {cliente_nome} não encontrado!")
        return

    user = cliente["user"]
    server = cliente["server"]
    port = cliente["port"]
    password = cliente["password"]
    path = cliente["path"]

    for arquivo in arquivos:
        if not os.path.isfile(arquivo):
            print(f"Arquivo {arquivo} não encontrado para {cliente_nome}.")
            continue

        # Enviar o arquivo via SCP (desativar verificação de chave do host)
        comando_scp = f"sshpass -p '{password}' scp -o StrictHostKeyChecking=no -P{port} {arquivo} {user}@{server}:{path}"
        print(f"Comando SCP: {comando_scp}")  # Exibir o comando SCP
        subprocess.run(comando_scp, shell=True, capture_output=False, check=True)

        # Executar o script remoto e capturar os logs no terminal (desativar verificação de chave do host)
        comando_ssh = (
            f"sshpass -p '{password}' ssh -o StrictHostKeyChecking=no -p{port} {user}@{server} "
            f"'sudo su - eacadm -c \"/u/saci/shells/importa_notas.sh\" 2>&1'"
        )
        print(f"Comando SSH: {comando_ssh}")  # Exibir o comando SSH
        subprocess.run(comando_ssh, shell=True, capture_output=False, check=True)

# Função principal para processar os clientes e enviar arquivos
def processar_clientes():
    config = carregar_configuracao()
    downloads_dir = "/home/sidney/Downloads"  # Diretório onde os arquivos estão

    for cliente_nome, dados_cliente in config["clientes"].items():
        if cliente_nome == "mueller":
            # Para Mueller, temos dois arquivos: FG e EL
            arquivos_excel = [os.path.join(downloads_dir, f"{cliente_nome}FG.xlsx"), os.path.join(downloads_dir, f"{cliente_nome}EL.xlsx")]
        elif cliente_nome == "muffato":
            # Para Muffato, temos dois arquivos: PR e SP
            arquivos_excel = [os.path.join(downloads_dir, f"{cliente_nome}PR.xlsx"), os.path.join(downloads_dir, f"{cliente_nome}SP.xlsx")]
        else:
            # Arquivo padrão
            arquivos_excel = [os.path.join(downloads_dir, f"{cliente_nome}.xlsx")]

        arquivos_csv = []
        
        # Verificar se o arquivo Excel existe para o cliente
        for xlsx_file in arquivos_excel:
            if os.path.isfile(xlsx_file):
                output_dir = os.path.join(downloads_dir, cliente_nome)

                # Criar o diretório de saída se não existir
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)

                # Converter o arquivo Excel para CSV
                csv_file = xlsx_to_csv(xlsx_file, output_dir)
                if csv_file:
                    arquivos_csv.append(csv_file)
            else:
                print(f"Arquivo {xlsx_file} não encontrado para {cliente_nome}.")
        
        # Enviar os arquivos convertidos
        if arquivos_csv:
            enviar_arquivo(cliente_nome, arquivos_csv)

# Executar o processo
if __name__ == "__main__":
    processar_clientes()
    print("Envio finalizado com sucesso!")